#!/usr/bin/env python3
import os
import sys
import json
import base64
import time
import re
from pathlib import Path
from datetime import datetime

# Vytvoření barev pro terminál pro prémiový inženýrský design (Sleek UI)
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

# Cesty k souborům (relativní k umístění skriptu pro plnou přenositelnost)
SCRIPT_DIR = Path(__file__).parent.resolve()
LISTINGS_PATH = SCRIPT_DIR / "bazos_active_listings.json"
CONFIG_PATH = SCRIPT_DIR / "bazos_config.json"
SESSION_STATE_PATH = SCRIPT_DIR / "bazos_session.json"


import atexit

class PlaywrightSessionManager:
    def __init__(self):
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

    def get_session(self):
        from datetime import datetime
        def log_psm(msg):
            try:
                with open("/tmp/thread_debug.log", "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] [PSM] {msg}\n")
            except Exception:
                pass

        is_active = False
        try:
            if self.browser and self.browser.is_connected() and self.page and not self.page.is_closed():
                _ = self.page.url
                is_active = True
        except Exception as e:
            log_psm(f"Session check error: {e}")
            is_active = False
            
        if not is_active:
            log_psm("Starting session reset/close")
            self.close()
            try:
                from playwright.sync_api import sync_playwright
            except ImportError:
                print(f"\n{Colors.WARNING}Instaluji knihovnu Playwright...{Colors.ENDC}")
                os.system(".venv/bin/pip install playwright")
                os.system(".venv/bin/playwright install chromium")
                from playwright.sync_api import sync_playwright
                
            log_psm("Calling sync_playwright().start()")
            self.playwright = sync_playwright().start()
            log_psm("sync_playwright().start() completed")
            try:
                log_psm("Launching browser (chrome)")
                self.browser = self.playwright.chromium.launch(channel="chrome", headless=False)
            except Exception as e:
                log_psm(f"Chrome launch failed: {e}. Launching default chromium.")
                self.browser = self.playwright.chromium.launch(headless=False)
            log_psm("Browser launched successfully")
            
            if SESSION_STATE_PATH.exists():
                log_psm("Loading session state (cookies)")
                self.context = self.browser.new_context(storage_state=str(SESSION_STATE_PATH))
            else:
                log_psm("Creating new context")
                self.context = self.browser.new_context()
            
            log_psm("Setting default timeout")
            self.context.set_default_timeout(30000)
            log_psm("Creating new page")
            self.page = self.context.new_page()
            log_psm("Session initialized successfully")
            
        return self.playwright, self.browser, self.context, self.page

    def save_state(self):
        if self.context:
            try:
                self.context.storage_state(path=str(SESSION_STATE_PATH))
            except Exception as e:
                print(f"  {Colors.WARNING}Nepodařilo se uložit stav relace: {e}{Colors.ENDC}")

    def close(self):
        if self.context:
            self.save_state()
        if self.browser:
            try:
                self.browser.close()
            except Exception:
                pass
        if self.playwright:
            try:
                self.playwright.stop()
            except Exception:
                pass
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

session_manager = PlaywrightSessionManager()
atexit.register(session_manager.close)



def load_data():
    try:
        if not LISTINGS_PATH.exists():
            with open(LISTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump({"active_listings": [], "sold_listings": []}, f, ensure_ascii=False, indent=2)
                
        with open(LISTINGS_PATH, "r", encoding="utf-8") as f:
            listings_data = json.load(f)
            
        if "active_listings" not in listings_data:
            listings_data["active_listings"] = []
        if "sold_listings" not in listings_data:
            listings_data["sold_listings"] = []
            
        if not CONFIG_PATH.exists():
            default_config = {
                "user": {
                    "email": "tuj_email@example.com",
                    "phone": "777123456",
                    "phone_verified": "+420777123456",
                    "default_ad_password_b64": "aGVzbG8xMjM=",
                    "name": "Tvoje Jméno",
                    "zip_code": "10000",
                    "location": "Praha 100 00"
                }
            }
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
                
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config_data = json.load(f)
            
        return listings_data, config_data.get("user", {})
    except Exception as e:
        print(f"{Colors.FAIL}Chyba při načítání databází: {e}{Colors.ENDC}")
        sys.exit(1)

def save_listings(data):
    try:
        with open(LISTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"{Colors.FAIL}Chyba při ukládání databáze: {e}{Colors.ENDC}")

# --- OneDrive & Excel Synchronizace ---
def sync_to_onedrive(data):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"{Colors.WARNING}Instaluji knihovnu openpyxl pro synchronizaci s OneDrivem...{Colors.ENDC}")
        os.system(".venv/bin/pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    # Detekce OneDrive cesty
    onedrive_dir = Path.home() / "Library/CloudStorage/OneDrive-Osobní"
    if not onedrive_dir.exists():
        onedrive_dir = Path.home() / "Desktop/Antigravity"
        
    onedrive_dir.mkdir(parents=True, exist_ok=True)
    file_path = onedrive_dir / "Inzerce - bazos.xlsx"
    
    wb = openpyxl.Workbook()
    
    # První list: Aktivní inzerce
    ws_active = wb.active
    ws_active.title = "Aktivní inzerce"
    
    # Druhý list: Historie prodejů
    ws_sold = wb.create_sheet(title="Historie prodejů")
    
    # Styly pro prémiový TERMS vzhled
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B2C82", end_color="4B2C82", fill_type="solid") # TERMS tmavě fialová
    title_font = Font(name=font_family, size=16, bold=True, color="4B2C82")
    data_font = Font(name=font_family, size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Pomocná funkce pro nastavení listu
    def setup_sheet(ws, sheet_title, listings, headers, is_sold=False):
        ws.views.sheetView[0].showGridLines = True
        
        # Titulní nadpis
        ws.merge_cells("A1:D1")
        ws["A1"] = sheet_title
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 30
        
        # Prázdný řádek 2
        ws.row_dimensions[2].height = 15
        
        # Hlavičky na řádku 3
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[3].height = 25
        
        current_row = 4
        for idx, ad in enumerate(listings, 1):
            ws.row_dimensions[current_row].height = 22
            row_fill = PatternFill(start_color="F8F6FC" if idx % 2 == 0 else "FFFFFF", end_color="F8F6FC" if idx % 2 == 0 else "FFFFFF", fill_type="solid")
            
            if not is_sold:
                # Výpočet stáří
                days_old = ad.get("days_old", 0)
                date_created = ad.get("date_created", "")
                if date_created and days_old == 0:
                    try:
                        dt = datetime.strptime(date_created, "%Y-%m-%d")
                        days_old = (datetime.today() - dt).days
                    except Exception:
                        pass
                
                # ["ID", "Název věci", "Inzerovaná cena (Kč)", "Datum vystavení", "Stáří (dní)", "Zhlédnutí", "Stav věci", "Stav", "Odkaz na Bazoš", "Složka s fotkami"]
                values = [
                    idx,
                    ad.get("title", ""),
                    ad.get("price", 0),
                    ad.get("date_created", ""),
                    days_old,
                    ad.get("views", 0),
                    ad.get("condition", "Nezadáno"),
                    ad.get("status", "Aktivní"),
                    ad.get("url", ""),
                    ad.get("local_photos_dir", "")
                ]
            else:
                # ["ID", "Název věci", "Původní cena (Kč)", "Prodejní cena (Kč)", "Rozdíl (Kč)", "Datum vystavení", "Datum prodeje", "Stav věci", "Poznámky"]
                price = ad.get("price", 0)
                sale_price = ad.get("sale_price", 0)
                diff = sale_price - price
                values = [
                    idx,
                    ad.get("title", ""),
                    price,
                    sale_price,
                    diff,
                    ad.get("date_created", ""),
                    ad.get("date_sold", ""),
                    ad.get("condition", "Nezadáno"),
                    ad.get("notes", "")
                ]
                
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                
                # Zarovnání a formátování podle sloupce
                if col_idx == 1:
                    cell.alignment = align_center
                elif not is_sold and col_idx == 3: # Cena active
                    cell.alignment = align_right
                    cell.number_format = '#,##0" Kč"'
                elif is_sold and col_idx in [3, 4, 5]: # Ceny sold
                    cell.alignment = align_right
                    cell.number_format = '#,##0" Kč"'
                elif not is_sold and col_idx in [4, 5, 6, 8]: # Datum, Stáří, Zhlédnutí, Stav active
                    cell.alignment = align_center
                elif is_sold and col_idx in [6, 7]: # Datumy sold
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
                # Zelený nebo červený badge pro stav aktivních inzerátů
                if not is_sold and col_idx == 8:
                    status_val = ad.get("status", "Aktivní")
                    if status_val == "Aktivní":
                        cell.fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, color="0F5132", bold=True)
                    elif status_val == "Expirováno":
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, color="842029", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, color="664D03", bold=True)
                    
                # Klikatelný hyperlink pro odkaz
                if not is_sold and col_idx == 9 and val:
                    cell.hyperlink = val
                    cell.font = Font(name=font_family, size=11, color="0D6EFD", underline="single")
                    
            current_row += 1
            
        # Přidání celkové sumy u prodejů
        if is_sold and len(listings) > 0:
            ws.row_dimensions[current_row].height = 24
            ws.cell(row=current_row, column=2).value = "Celkem prodáno"
            ws.cell(row=current_row, column=2).font = Font(name=font_family, size=11, bold=True)
            ws.cell(row=current_row, column=2).alignment = align_right
            
            # Součet původních cen
            cell_orig = ws.cell(row=current_row, column=3)
            cell_orig.value = f"=SUM(C4:C{current_row-1})"
            cell_orig.font = Font(name=font_family, size=11, bold=True)
            cell_orig.alignment = align_right
            cell_orig.number_format = '#,##0" Kč"'
            cell_orig.border = thin_border
            
            # Součet prodejních cen
            cell_sale = ws.cell(row=current_row, column=4)
            cell_sale.value = f"=SUM(D4:D{current_row-1})"
            cell_sale.font = Font(name=font_family, size=11, bold=True, color="0F5132")
            cell_sale.alignment = align_right
            cell_sale.number_format = '#,##0" Kč"'
            cell_sale.border = thin_border
            cell_sale.fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
            
            # Součet rozdílů
            cell_diff = ws.cell(row=current_row, column=5)
            cell_diff.value = f"=SUM(E4:E{current_row-1})"
            cell_diff.font = Font(name=font_family, size=11, bold=True)
            cell_diff.alignment = align_right
            cell_diff.number_format = '#,##0" Kč"'
            cell_diff.border = thin_border
            
        # Automatická šířka sloupců
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    setup_sheet(ws_active, "Aktivní inzerce na Bazoši", data.get("active_listings", []), ["ID", "Název věci", "Inzerovaná cena (Kč)", "Datum vystavení", "Stáří (dní)", "Zhlédnutí", "Stav věci", "Stav", "Odkaz na Bazoš", "Složka s fotkami"], is_sold=False)
    setup_sheet(ws_sold, "Historie prodejů", data.get("sold_listings", []), ["ID", "Název věci", "Původní cena (Kč)", "Prodejní cena (Kč)", "Rozdíl (Kč)", "Datum vystavení", "Datum prodeje", "Stav věci", "Poznámky"], is_sold=True)
    
    wb.save(file_path)
    print(f"{Colors.GREEN}✓ Tabulka Excel úspěšně synchronizována na OneDrive: {file_path}{Colors.ENDC}")

# --- Pomocné funkce pro parsování URL a domén ---
def extract_ad_id(url):
    if not url:
        return None
    match = re.search(r'/inzerat/(\d+)', url)
    if match:
        return match.group(1)
    return None

def extract_subdomain(url):
    if not url:
        return "dum.bazos.cz"
    match = re.search(r'https://([^/]+)', url)
    if match:
        return match.group(1)
    return "dum.bazos.cz"

def get_target_domain(title, original_url=""):
    if original_url and "nabytek" in original_url:
        return "nabytek.bazos.cz"
    # Fallback podle klíčových slov
    nabytek_keywords = ["stůl", "židle", "skříň", "komoda", "postel", "matrace", "sedačka", "pohovka", "křeslo", "stoly", "židle", "nabytek", "jídelní", "sedák"]
    title_lower = title.lower()
    if any(kw in title_lower for kw in nabytek_keywords):
        return "nabytek.bazos.cz"
    return "dum.bazos.cz"

# --- Zobrazení terminálové tabulky (UX) ---
def display_listings_summary(data):
    try:
        from tabulate import tabulate
    except ImportError:
        os.system(".venv/bin/pip install tabulate")
        from tabulate import tabulate

    active = data.get("active_listings", [])
    sold = data.get("sold_listings", [])
    
    print(f"\n{Colors.CYAN}{Colors.BOLD}📊 AKTUÁLNÍ PŘEHLED AKTIVNÍ INZERCE{Colors.ENDC}")
    if not active:
        print("  Žádné aktivní inzeráty.")
    else:
        table_active = []
        for idx, ad in enumerate(active, 1):
            price_str = f"{ad['price']:,} Kč".replace(',', ' ') if ad['price'] > 0 else "Zdarma"
            url_short = ad.get("url", "")[:35] + "..." if len(ad.get("url", "")) > 35 else ad.get("url", "Není odkaz")
            
            # Formátování stavu
            status = ad.get('status', 'Aktivní')
            if status == "Expirováno":
                status_formatted = f"{Colors.FAIL}Expirováno{Colors.ENDC}"
            elif status == "Aktivní":
                status_formatted = f"{Colors.GREEN}Aktivní{Colors.ENDC}"
            else:
                status_formatted = f"{Colors.BLUE}{status}{Colors.ENDC}"
                
            # Výpočet stáří
            days_old = ad.get("days_old", 0)
            date_created = ad.get("date_created", "")
            if date_created and days_old == 0:
                try:
                    dt = datetime.strptime(date_created, "%Y-%m-%d")
                    days_old = (datetime.today() - dt).days
                except Exception:
                    pass
            days_old_str = f"{days_old} dní" if days_old >= 0 else "Nezadáno"
            
            views = ad.get("views", 0)
            views_str = f"{views}x"
            
            table_active.append([
                idx,
                ad['title'][:35],
                price_str,
                ad.get('date_created', 'Nezadáno'),
                days_old_str,
                views_str,
                ad.get('condition', 'Nezadáno')[:20],
                status_formatted,
                url_short
            ])
        print(tabulate(table_active, headers=["ID", "Název věci", "Cena", "Datum", "Stáří", "Zhlédnutí", "Stav věci", "Stav", "Odkaz"], tablefmt="fancy_grid"))

    print(f"\n{Colors.BLUE}{Colors.BOLD}📜 HISTORIE PRODANÝCH VĚCÍ{Colors.ENDC}")
    if not sold:
        print("  Žádná historie prodejů.")
    else:
        table_sold = []
        total_profit = 0
        for idx, ad in enumerate(sold, 1):
            orig_price_str = f"{ad['price']:,} Kč".replace(',', ' ')
            sale_price = ad.get('sale_price', ad['price'])
            sale_price_str = f"{sale_price:,} Kč".replace(',', ' ')
            diff = sale_price - ad['price']
            diff_color = Colors.GREEN if diff >= 0 else Colors.FAIL
            diff_str = f"{diff_color}{diff:+,} Kč{Colors.ENDC}".replace(',', ' ') if diff != 0 else "0 Kč"
            total_profit += sale_price
            table_sold.append([
                idx,
                ad['title'][:40],
                orig_price_str,
                sale_price_str,
                diff_str,
                ad.get('date_sold', 'Nezadáno'),
                ad.get('notes', 'Bez poznámky')[:35]
            ])
        print(tabulate(table_sold, headers=["ID", "Název věci", "Pův. cena", "Prodejní", "Rozdíl", "Datum prodeje", "Poznámka"], tablefmt="fancy_grid"))
        print(f"💰 {Colors.BOLD}Celková hodnota prodaných věcí:{Colors.ENDC} {Colors.GREEN}{total_profit:,} Kč{Colors.ENDC}".replace(',', ' '))

# --- Aktualizace stavů z Bazoše na pozadí ---
def parse_bazos_date(date_text):
    """
    Převede české relativní nebo absolutní datum z Bazoše do formátu YYYY-MM-DD.
    Příklady vstupů: '[29.6. 2026]', '[Dnes - 12:34]', '[Včera]', '29.6. 2026'
    """
    from datetime import datetime, timedelta
    try:
        # Odstraníme hranaté závorky
        cleaned = date_text.replace("[", "").replace("]", "").strip()
        cleaned_lower = cleaned.lower()
        
        if "dnes" in cleaned_lower:
            return datetime.today().strftime("%Y-%m-%d")
        elif "včera" in cleaned_lower:
            return (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        elif "předevčírem" in cleaned_lower:
            return (datetime.today() - timedelta(days=2)).strftime("%Y-%m-%d")
            
        # Standardní české datum, např. "29.6. 2026" nebo "29. 6. 2026"
        # Odstraníme mezery za tečkami
        cleaned = re.sub(r'\s+', '', cleaned) # "29.6.2026"
        
        # Extrahujeme pouze "den.měsíc.rok"
        match = re.search(r'(\d+)\.(\d+)\.(\d+)', cleaned)
        if match:
            day = int(match.group(1))
            month = int(match.group(2))
            year = int(match.group(3))
            parsed_date = datetime(year, month, day)
            return parsed_date.strftime("%Y-%m-%d")
    except Exception:
        pass
    # Fallback na dnešní datum, pokud se nepodaří parsovat
    return datetime.today().strftime("%Y-%m-%d")

# --- Aktualizace stavů z Bazoše na pozadí ---
def cli_update_listings_from_bazos(data, is_web=False):
    from bs4 import BeautifulSoup
    import re
    
    # Načteme uživatelský config pro e-mail a telefon
    _, user_config = load_data()
    
    email_val = user_config.get("email", "").strip()
    phone_val = user_config.get("phone", "").strip()
    if not email_val or not phone_val or len(phone_val) != 9 or not phone_val.isdigit():
        raise Exception("V konfiguraci chybí platné 9-místné telefonní číslo pro SMS ověření.")
    
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print(f"\n{Colors.WARNING}Instaluji knihovnu Playwright...{Colors.ENDC}")
        os.system(".venv/bin/pip install playwright")
        os.system(".venv/bin/playwright install chromium")
        from playwright.sync_api import sync_playwright
        
    print(f"\n{Colors.HEADER}{Colors.BOLD}🔄 AKTUALIZACE STAVŮ INZERÁTŮ Z BAZOŠE (Moje inzeráty){Colors.ENDC}")
    print(f"{Colors.BLUE}Spouštím/připojuji prohlížeč pro přihlášení a stažení inzerátů...{Colors.ENDC}\n")
    
    html_content = ""
    from contextlib import nullcontext
    with nullcontext():
        try:
            p, browser, context, page = session_manager.get_session()
        except Exception as e:
            print(f"{Colors.FAIL}Chyba při inicializaci prohlížeče: {e}{Colors.ENDC}")
            return
            
        # Přejdeme na Moje inzeráty
        page.goto("https://www.bazos.cz/moje-inzeraty.php")
        
        # Zkontrolujeme, zda je zobrazen formulář pro přihlášení
        try:
            email_input = page.locator("input[name='mail']")
            phone_input = page.locator("input[name='telefon']")
            
            if email_input.is_visible(timeout=3000):
                email_val = user_config.get("email", "tuj_email@example.com")
                phone_val = user_config.get("phone", "777123456")
                
                print(f"  {Colors.BLUE}Relace neexistuje nebo vypršela. Vyplňuji přihlašovací údaje...{Colors.ENDC}")
                email_input.fill(email_val)
                phone_input.fill(phone_val)
                
                # Zkontrolujeme, zda je telefon již ověřený (tzn. je vidět tlačítko Vypsat inzeráty)
                list_btn = page.locator("input[type='submit'][value='Vypsat inzeráty']")
                if list_btn.is_visible(timeout=2000):
                    print(f"  {Colors.GREEN}✓ Telefon je již ověřený, klikám na Vypsat inzeráty...{Colors.ENDC}")
                    list_btn.click()
                    time.sleep(2)
                else:
                    # Klikneme na Ověřit
                    submit_btn = page.locator("input[type='submit'][value='Ověřit']")
                    submit_btn.click()
                    print(f"  {Colors.GREEN}✓ Odeslán požadavek na SMS kód.{Colors.ENDC}")
                    
                    # Nyní čekáme, až se objeví pole pro SMS kód kodd
                    code_input = page.locator("input[name='kodd']")
                    code_input.wait_for(timeout=10000)
                
                if is_web:
                    print(f"  {Colors.WARNING}💬 [WEB] Přihlášení vyžaduje SMS kód. Zadej ho prosím přímo v otevřeném okně prohlížeče a klikni na 'Vypsat inzeráty'...{Colors.ENDC}")
                    try:
                        # Čekáme na zmizení formuláře (tzn. úspěšné přihlášení a přesměrování) po dobu až 60 sekund
                        page.wait_for_selector("input[name='kodd']", state="hidden", timeout=60000)
                        time.sleep(2)
                    except Exception:
                        print(f"  {Colors.WARNING}Vypršel časový limit pro zadání SMS kódu v prohlížeči.{Colors.ENDC}")
                else:
                    # Požádáme uživatele v CLI o zadání SMS kódu
                    print(f"\n{Colors.BOLD}{Colors.HEADER}💬 SMS OVĚŘENÍ BAZOŠE:{Colors.ENDC}")
                    print(f"{Colors.BOLD}Na tvůj mobilní telefon ({phone_val}) byl odeslán SMS kód.{Colors.ENDC}")
                    sms_code = input(f"{Colors.BOLD}Zadej 6místný SMS kód: {Colors.ENDC}").strip()
                    if not sms_code:
                        print(f"{Colors.FAIL}SMS kód nebyl zadán. Ruším synchronizaci.{Colors.ENDC}")
                        session_manager.close()
                        return
                    code_input.fill(sms_code)
                    # Klikneme na odeslat - Vypsat inzeráty
                    list_btn = page.locator("input[type='submit'][value='Vypsat inzeráty']")
                    list_btn.click()
                    time.sleep(2)
            else:
                print(f"  {Colors.GREEN}✓ Úspěšně přihlášeno pomocí uložené relace.{Colors.ENDC}")
        except Exception as login_err:
            print(f"  {Colors.WARNING}Přihlašovací formulář se neobjevil nebo nastala chyba: {login_err}{Colors.ENDC}")
            if is_web:
                raise Exception(f"Nepodařilo se přihlásit k Bazoši: {login_err}")
            print(f"  {Colors.BLUE}Zkouším rovnou načíst přehled inzerátů...{Colors.ENDC}")
            
        try:
            # Počkáme chvíli na vykreslení stránky
            page.wait_for_load_state("networkidle", timeout=5000)
        except Exception:
            pass
            
        # Uložíme/aktualizujeme platnou relaci, pokud jsme přihlášeni
        try:
            if not page.locator("input[name='mail']").is_visible(timeout=2000):
                session_manager.save_state()
                print(f"  {Colors.GREEN}✓ Relace uložena/aktualizována pro příště.{Colors.ENDC}")
        except Exception as state_err:
            print(f"  {Colors.WARNING}Nepodařilo se uložit stav relace: {state_err}{Colors.ENDC}")
            
        html_content = page.content()
        pass


        
    if not html_content:
        print(f"{Colors.FAIL}Nepodařilo se stáhnout obsah stránky Bazoše!{Colors.ENDC}")
        return
        
    soup = BeautifulSoup(html_content, "html.parser")
    
    # Hledáme všechny inzeráty
    ad_elements = soup.find_all(class_=re.compile(r"\binzeraty\b"))
    
    scraped_listings = []
    print(f"\n{Colors.BLUE}Nalezeno {len(ad_elements)} inzerátů na Bazoši k vyparsování...{Colors.ENDC}")
    
    for el in ad_elements:
        try:
            # Nadpis a URL
            nadpis_el = el.find(class_="nadpis")
            if not nadpis_el:
                continue
            a_tag = nadpis_el.find("a")
            if not a_tag:
                continue
                
            title_text = a_tag.get_text().strip()
            ad_url = a_tag.get("href", "")
            
            # Pokud je URL relativní, doplníme doménu
            if ad_url.startswith("/"):
                ad_url = "https://www.bazos.cz" + ad_url
            elif not ad_url.startswith("http"):
                ad_url = "https://" + ad_url
                
            # Cena
            cena_el = el.find(class_="inzeratycena")
            price_val = 0
            if cena_el:
                cena_text = cena_el.get_text().replace(" ", "").replace("\xa0", "")
                price_match = re.search(r"(\d+)", cena_text)
                if price_match:
                    price_val = int(price_match.group(1))
                    
            # Zhlédnutí
            views_el = el.find(class_="inzeratyview")
            views_val = 0
            if views_el:
                views_text = views_el.get_text().replace(" ", "").replace("\xa0", "")
                views_match = re.search(r"(\d+)", views_text)
                if views_match:
                    views_val = int(views_match.group(1))
                    
            # Stáří / Datum vytvoření
            date_el = el.find(class_="velikost10")
            date_str = ""
            if date_el:
                date_text = date_el.get_text().strip()
                date_str = parse_bazos_date(date_text)
                
            scraped_listings.append({
                "title": title_text,
                "url": ad_url,
                "price": price_val,
                "views": views_val,
                "date_created": date_str
            })
        except Exception:
            continue
            
    # Synchronizace s lokální databází
    local_active = data.get("active_listings", [])
    
    updated_count = 0
    imported_count = 0
    expired_count = 0
    
    new_active_listings = []
    matched_scraped_indices = set()
    
    # 1. Nejprve spárujeme stávající lokální inzeráty
    for local_ad in local_active:
        best_scraped_match = None
        best_scraped_idx = -1
        
        local_title_50 = local_ad["title"][:50].lower().strip()
        local_url = local_ad.get("url", "").strip()
        
        # Zkusíme přesný match podle URL
        if local_url:
            for s_idx, scraped_ad in enumerate(scraped_listings):
                if s_idx in matched_scraped_indices:
                    continue
                if scraped_ad["url"].strip() == local_url:
                    best_scraped_match = scraped_ad
                    best_scraped_idx = s_idx
                    break
                    
        # Pokud nenašel, zkusíme match podle oříznutého Nadpisu (50 znaků)
        if not best_scraped_match:
            for s_idx, scraped_ad in enumerate(scraped_listings):
                if s_idx in matched_scraped_indices:
                    continue
                scraped_title_50 = scraped_ad["title"][:50].lower().strip()
                if scraped_title_50 == local_title_50:
                    best_scraped_match = scraped_ad
                    best_scraped_idx = s_idx
                    break
                    
        if best_scraped_match:
            matched_scraped_indices.add(best_scraped_idx)
            # Aktualizujeme lokální inzerát
            local_ad["url"] = best_scraped_match["url"]
            local_ad["price"] = best_scraped_match["price"]
            local_ad["views"] = best_scraped_match["views"]
            local_ad["date_created"] = best_scraped_match["date_created"]
            local_ad["status"] = "Aktivní"
            
            # Přepočet stáří
            try:
                dt = datetime.strptime(best_scraped_match["date_created"], "%Y-%m-%d")
                local_ad["days_old"] = (datetime.today() - dt).days
            except Exception:
                local_ad["days_old"] = 0
                
            print(f"  {Colors.GREEN}✓ Aktualizováno:{Colors.ENDC} '{local_ad['title']}' -> Zhlédnutí: {local_ad['views']}, Cena: {local_ad['price']} Kč, Stáří: {local_ad['days_old']} dní")
            updated_count += 1
            new_active_listings.append(local_ad)
        else:
            # Inzerát na Bazoši chybí -> Expiroval
            local_ad["status"] = "Expirováno"
            print(f"  {Colors.FAIL}✗ Expirováno / Smazáno na Bazoši:{Colors.ENDC} '{local_ad['title']}'")
            expired_count += 1
            new_active_listings.append(local_ad)
            
    # 2. Automaticky importujeme nově nalezené inzeráty z Bazoše, které v JSONu nemáme
    for s_idx, scraped_ad in enumerate(scraped_listings):
        if s_idx in matched_scraped_indices:
            continue
            
        default_pwd_b64 = user_config.get("default_ad_password_b64", "aGVzbG8xMjM=")
        
        # Generování cesty pro fotky bez závislostí
        import unicodedata
        def simple_slugify(text):
            text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
            text = text.lower()
            text = re.sub(r'[^a-z0-9\s_]', '', text)
            text = re.sub(r'[\s_]+', '_', text).strip('_')
            return text
            
        folder_name = simple_slugify(scraped_ad["title"])
        photos_dir = str(Path.home() / "Desktop" / "Antigravity" / "Bazos_Photos" / folder_name)
        os.makedirs(photos_dir, exist_ok=True)
        
        days_old_val = 0
        try:
            dt = datetime.strptime(scraped_ad["date_created"], "%Y-%m-%d")
            days_old_val = (datetime.today() - dt).days
        except Exception:
            pass
            
        new_ad = {
            "title": scraped_ad["title"],
            "price": scraped_ad["price"],
            "ad_password_b64": default_pwd_b64,
            "date_created": scraped_ad["date_created"],
            "days_old": days_old_val,
            "location": user_config.get("location", "Praha 100 00"),
            "views": scraped_ad["views"],
            "url": scraped_ad["url"],
            "local_photos_dir": photos_dir,
            "description": "Automaticky importovaný inzerát z Bazoše. Doplňte prosím popis.",
            "bookmarklet_uri": "",
            "condition": "Aktivní",
            "status": "Aktivní",
            "notes": "Automatický import"
        }
        
        print(f"  {Colors.HEADER}➕ Nově importováno z Bazoše:{Colors.ENDC} '{new_ad['title']}' (Cena: {new_ad['price']} Kč, URL: {new_ad['url']})")
        imported_count += 1
        new_active_listings.append(new_ad)
        
    data["active_listings"] = new_active_listings
    save_listings(data)
    
    print(f"\n{Colors.GREEN}✓ Synchronizace s Bazošem úspěšně dokončena!{Colors.ENDC}")
    print(f"  Aktualizováno: {Colors.BOLD}{updated_count}{Colors.ENDC}")
    print(f"  Expirováno: {Colors.BOLD}{expired_count}{Colors.ENDC}")
    print(f"  Nově importováno: {Colors.BOLD}{imported_count}{Colors.ENDC}")
    
    sync_to_onedrive(data)

# --- CLI Průvodce pro přidání nové věci ---
def cli_add_listing(data, user_config):
    print(f"\n{Colors.HEADER}{Colors.BOLD}🆕 PRŮVODCE PŘIDÁNÍM NOVÉ VĚCI K INZERCI{Colors.ENDC}")
    title = input(f"{Colors.BOLD}Název věci (max 50 znaků): {Colors.ENDC}").strip()
    while not title or len(title) > 50:
        if len(title) > 50:
            print(f"{Colors.FAIL}Název je příliš dlouhý ({len(title)} znaků). Musí mít max 50 znaků!{Colors.ENDC}")
        title = input(f"{Colors.BOLD}Zadej název věci (max 50 znaků): {Colors.ENDC}").strip()
        
    try:
        price = int(input(f"{Colors.BOLD}Inzerovaná cena v Kč (pouze číslo): {Colors.ENDC}").strip())
    except ValueError:
        price = 0
        print(f"{Colors.WARNING}Neplatná cena, nastaveno na 0 Kč (Zdarma).{Colors.ENDC}")
        
    print(f"\n{Colors.BLUE}Zadej popis inzerátu (stiskni dvakrát Enter pro dokončení):{Colors.ENDC}")
    desc_lines = []
    while True:
        line = input()
        if line == "" and desc_lines and desc_lines[-1] == "":
            break
        desc_lines.append(line)
    description = "\n".join(desc_lines).strip()
    
    condition = input(f"{Colors.BOLD}Stav věci (např. Skvělý stav / Použitý): {Colors.ENDC}").strip()
    if not condition:
        condition = "Dobrý stav, plně funkční"
        
    default_dir_name = re.sub(r'[^a-z0-9]', '_', title.lower())
    default_dir_name = re.sub(r'_+', '_', default_dir_name).strip('_')
    suggested_photos_dir = str(Path.home() / "Desktop" / "Antigravity" / "Bazos_Photos" / default_dir_name)
    
    photos_dir = input(f"{Colors.BOLD}Složka s fotkami [výchozí: {suggested_photos_dir}]: {Colors.ENDC}").strip()
    if not photos_dir:
        photos_dir = suggested_photos_dir
        
    # Vytvoření složky pro fotky, pokud neexistuje
    os.makedirs(photos_dir, exist_ok=True)
    print(f"{Colors.GREEN}Složka s fotkami připravena: {photos_dir}{Colors.ENDC}")
    print(f"{Colors.CYAN}Tip: Do této složky nakopíruj fotky (foto_1.jpg, foto_2.jpg atd. nebo jakékoliv jiné).{Colors.ENDC}")
    
    # Zakódování výchozího hesla
    password_b64 = user_config.get("default_ad_password_b64", "aGVzbG8xMjM=")
    
    new_ad = {
        "title": title,
        "price": price,
        "ad_password_b64": password_b64,
        "date_created": datetime.today().strftime('%Y-%m-%d'),
        "days_old": 0,
        "location": user_config.get("location", "Praha 100 00"),
        "views": 0,
        "url": "",
        "local_photos_dir": photos_dir,
        "description": description,
        "condition": condition,
        "status": "Aktivní",
        "notes": ""
    }
    
    data["active_listings"].append(new_ad)
    save_listings(data)
    print(f"\n{Colors.GREEN}✓ Věc '{title}' úspěšně přidána do databáze!{Colors.ENDC}")
    sync_to_onedrive(data)

# --- Zaznamenání prodeje věci ---
def cli_record_sale(data):
    active = data.get("active_listings", [])
    if not active:
        print(f"{Colors.FAIL}Nemáš žádné aktivní inzeráty k prodeji!{Colors.ENDC}")
        return
        
    print(f"\n{Colors.HEADER}{Colors.BOLD}💰 ZAZNAMENAT PRODEJ VĚCI{Colors.ENDC}")
    for idx, ad in enumerate(active, 1):
        print(f"  [{idx}] {ad['title']} (Inzerovaná cena: {ad['price']} Kč)")
        
    choice = input(f"\nZadej číslo věci, kterou jsi prodal (1-{len(active)}): ").strip()
    try:
        idx = int(choice) - 1
        if idx < 0 or idx >= len(active):
            raise ValueError
        selected_ad = active[idx]
    except ValueError:
        print(f"{Colors.FAIL}Neplatná volba!{Colors.ENDC}")
        return
        
    try:
        sale_price = int(input(f"Skutečná prodejní cena v Kč [výchozí: {selected_ad['price']} Kč]: ").strip() or selected_ad['price'])
    except ValueError:
        sale_price = selected_ad['price']
        
    notes = input("Poznámka k prodeji (např. kdo koupil, slevy atd.): ").strip()
    
    # Přesun z aktivních do prodaných
    selected_ad["status"] = "Prodáno"
    selected_ad["sale_price"] = sale_price
    selected_ad["date_sold"] = datetime.today().strftime('%Y-%m-%d')
    selected_ad["notes"] = notes
    
    data["sold_listings"].append(selected_ad)
    active.pop(idx)
    
    save_listings(data)
    print(f"\n{Colors.GREEN}✓ Prodej úspěšně zaznamenán! Věc byla přesunuta do historie.{Colors.ENDC}")
    sync_to_onedrive(data)

# --- Změna ceny ---
def cli_change_price(data, user_config, page_runner):
    active = data.get("active_listings", [])
    if not active:
        print(f"{Colors.FAIL}Nemáš žádné aktivní inzeráty pro změnu ceny!{Colors.ENDC}")
        return
        
    print(f"\n{Colors.HEADER}{Colors.BOLD}🏷️  ZMĚNIT CENU INZERÁTU{Colors.ENDC}")
    for idx, ad in enumerate(active, 1):
        print(f"  [{idx}] {ad['title']} (Aktuální cena: {ad['price']} Kč)")
        
    choice = input(f"\nZadej číslo inzerátu pro změnu ceny (1-{len(active)}): ").strip()
    try:
        idx = int(choice) - 1
        if idx < 0 or idx >= len(active):
            raise ValueError
        selected_ad = active[idx]
    except ValueError:
        print(f"{Colors.FAIL}Neplatná volba!{Colors.ENDC}")
        return
        
    try:
        new_price = int(input(f"Zadej novou cenu v Kč pro '{selected_ad['title']}': ").strip())
    except ValueError:
        print(f"{Colors.FAIL}Cena musí být číslo!{Colors.ENDC}")
        return
        
    selected_ad["price"] = new_price
    save_listings(data)
    print(f"{Colors.GREEN}✓ Cena byla místně aktualizována v databázi na {new_price} Kč.{Colors.ENDC}")
    sync_to_onedrive(data)
    
    # Ověříme, zda má inzerát URL adresu na Bazoši
    url = selected_ad.get("url")
    if not url:
        print(f"{Colors.WARNING}Tento inzerát nemá uloženou URL adresu na Bazoši (ještě nebyl vystaven). Změna se promítne při příštím vystavení.{Colors.ENDC}")
        return
        
    sync_online = input(f"\n{Colors.BOLD}Chceš novou cenu promítnout ihned přímo na Bazoš? [y/N]: {Colors.ENDC}").strip().lower()
    if sync_online == 'y':
        print(f"\n{Colors.BLUE}Spouštím poloautomatickou změnu ceny přes Playwright...{Colors.ENDC}")
        page_runner(selected_ad, user_config, action="edit_price", extra_val=str(new_price))

# --- Poloautomatická správa inzerátu přes Playwright (Vystavení, Smazání, Editace ceny) ---
def _run_playwright_action_impl(ad, user_config, action="post", extra_val=None, is_web=False):
    email_val = user_config.get("email", "").strip()
    phone_val = user_config.get("phone", "").strip()
    if not email_val or not phone_val or len(phone_val) != 9 or not phone_val.isdigit():
        raise Exception("V konfiguraci chybí platné 9-místné telefonní číslo pro SMS ověření.")
        
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print(f"\n{Colors.WARNING}Instaluji knihovnu Playwright...{Colors.ENDC}")
        os.system(".venv/bin/pip install playwright")
        os.system(".venv/bin/playwright install chromium")
        from playwright.sync_api import sync_playwright

    title = ad["title"]
    description = ad["description"]
    price = str(ad["price"]) if not extra_val else extra_val
    photos_dir = ad["local_photos_dir"]
    url = ad.get("url", "")
    
    ad_id = extract_ad_id(url)
    subdomain = extract_subdomain(url)
    password_b64 = ad.get("ad_password_b64", "aGVzbG8xMjM=")
    password = base64.b64decode(password_b64).decode("utf-8")
    
    # Zkrácení nadpisu na limit Bazoše
    if len(title) > 50:
        title = title[:50]

    photos = []
    if os.path.exists(photos_dir):
        raw_files = os.listdir(photos_dir)
        jpg_files = [f for f in raw_files if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        jpg_files = sorted(jpg_files, key=lambda x: (not x.startswith("foto_"), x))
        photos = [os.path.join(photos_dir, f) for f in jpg_files]

    from contextlib import nullcontext
    with nullcontext():
        try:
            p, browser, context, page = session_manager.get_session()
        except Exception as e:
            print(f"{Colors.FAIL}Chyba při inicializaci prohlížeče: {e}{Colors.ENDC}")
            return False
        
        # --- AKCE: SMAZÁNÍ (DELETE) ---
        if action == "delete":
            if not ad_id:
                print(f"{Colors.FAIL}Chyba: Chybí ID inzerátu pro smazání!{Colors.ENDC}")
                return False
                
            print(f"\n{Colors.BLUE}Směřuji na mazací formulář: https://{subdomain}/delete.php?id={ad_id}...{Colors.ENDC}")
            page.goto(f"https://{subdomain}/delete.php?id={ad_id}")
            
            try:
                # Vyplníme heslo
                password_input = page.locator("input[name='heslo'], input[type='password']")
                password_input.wait_for(timeout=5000)
                password_input.fill(password)
                print(f"  {Colors.GREEN}✓ Heslo inzerátu předvyplněno.{Colors.ENDC}")
                
                # Zaškrtneme Smazat (pokud jsou radio buttons)
                radio_delete = page.locator("input[type='radio'][value='delete'], input[value='2']")
                if radio_delete.count() > 0:
                    radio_delete.click()
                    print(f"  {Colors.GREEN}✓ Možnost 'Smazat' vybrána.{Colors.ENDC}")
                
                # Klikneme na odeslat/potvrdit
                submit_btn = page.locator("input[type='submit'], button[type='submit']")
                submit_btn.first.click()
                print(f"\n{Colors.GREEN}✓ Formulář odeslán.{Colors.ENDC}")
                print(f"{Colors.BOLD}👉 Dokonči smazání v prohlížeči (např. výběr důvodu)...{Colors.ENDC}")
                if not is_web:
                    print("stiskni [Enter] v terminálu pro pokračování...")
                    input()
                try:
                    context.storage_state(path=str(SESSION_STATE_PATH))
                except Exception:
                    pass
                return True
            except Exception as delete_err:
                print(f"{Colors.FAIL}Chyba při mazání inzerátu: {delete_err}{Colors.ENDC}")
                print("Dokonči prosím smazání ručně v otevřeném prohlížeči...")
                if not is_web:
                    print("stiskni [Enter] v terminálu pro pokračování...")
                    input()
                try:
                    context.storage_state(path=str(SESSION_STATE_PATH))
                except Exception:
                    pass
                return True

        # --- AKCE: EDITACE CENY (EDIT PRICE) ---
        elif action == "edit_price":
            if not ad_id:
                print(f"{Colors.FAIL}Chyba: Chybí ID inzerátu pro editaci!{Colors.ENDC}")
                return False
                
            print(f"\n{Colors.BLUE}Směřuji na administrační formulář: https://{subdomain}/delete.php?id={ad_id}...{Colors.ENDC}")
            page.goto(f"https://{subdomain}/delete.php?id={ad_id}")
            
            try:
                # Vyplníme heslo
                password_input = page.locator("input[name='heslo'], input[type='password']")
                password_input.wait_for(timeout=5000)
                password_input.fill(password)
                
                # Vybereme editaci
                radio_edit = page.locator("input[type='radio'][value='edit'], input[value='1']")
                if radio_edit.count() > 0:
                    radio_edit.click()
                    print(f"  {Colors.GREEN}✓ Vybrána editace inzerátu.{Colors.ENDC}")
                    
                submit_btn = page.locator("input[type='submit'], button[type='submit']")
                submit_btn.first.click()
                
                # Čekáme na načtení editačního formuláře ceny
                price_input = page.locator("input[name='cena'], #cena")
                price_input.wait_for(timeout=5000)
                price_input.fill(price)
                
                print(f"\n{Colors.GREEN}🎉 Nová cena {price} Kč byla úspěšně předvyplněna na Bazoši!{Colors.ENDC}")
                print(f"{Colors.BOLD}👉 Zkontroluj inzerát v prohlížeči a klikni dole na 'Upravit' pro uložení změn.{Colors.ENDC}")
                if not is_web:
                    print("Po dokončení stiskni [Enter] zde v terminálu...")
                    input()
                try:
                    context.storage_state(path=str(SESSION_STATE_PATH))
                except Exception:
                    pass
                return True
            except Exception as edit_err:
                print(f"{Colors.FAIL}Nepodařilo se plně automatizovat změnu ceny: {edit_err}{Colors.ENDC}")
                print("Uprav prosím cenu ručně v otevřeném prohlížeči a ulož ji.")
                if not is_web:
                    print("Poté stiskni [Enter] v terminálu...")
                    input()
                try:
                    context.storage_state(path=str(SESSION_STATE_PATH))
                except Exception:
                    pass
                return True

        # --- AKCE: POST / VYSTAVENÍ (DEFAULT) ---
        else:
            target_domain = get_target_domain(title, url)
            print(f"\n{Colors.BLUE}Směřuji na přidání inzerátu: https://{target_domain}/pridat-inzerat.php{Colors.ENDC}")
            page.goto(f"https://{target_domain}/pridat-inzerat.php")
            
            def find_and_fill(label, selectors, value):
                for selector in selectors:
                    try:
                        loc = page.locator(selector)
                        if loc.is_visible(timeout=500):
                            loc.fill(value)
                            print(f"  {Colors.GREEN}✓ {label}{Colors.ENDC} vyplněno (selektor '{selector}')")
                            return True
                    except Exception:
                        continue
                return False

            def select_rubrika_first():
                try:
                    selector = "select[name='rubrikyvybrat']"
                    select_loc = page.locator(selector)
                    if select_loc.count() > 0 and select_loc.is_visible(timeout=500):
                        options_elements = select_loc.locator("option").all()
                        best_value = None
                        best_score = -1
                        best_label = ""
                        
                        title_lower = title.lower()
                        description_lower = description.lower()
                        
                        for opt in options_elements:
                            val = opt.get_attribute("value")
                            if not val or val == "" or val == "0":
                                continue
                            label = opt.inner_text().strip().lower()
                            
                            score = 0
                            if label in title_lower:
                                score += 10
                            if label in description_lower:
                                score += 2
                            
                            if score > best_score:
                                best_score = score
                                best_value = val
                                best_label = opt.inner_text().strip()
                        
                        if best_value and best_score > 0:
                            current_value = select_loc.evaluate("el => el.value")
                            if current_value != best_value:
                                print(f"  {Colors.BLUE}Změna rubriky na '{best_label}'...{Colors.ENDC}")
                                select_loc.select_option(value=best_value)
                                try:
                                    page.wait_for_load_state("networkidle", timeout=3000)
                                except Exception:
                                    pass
                                time.sleep(1.5)
                                print(f"  {Colors.GREEN}✓ Rubrika změněna na: '{best_label}'{Colors.ENDC}")
                                return True
                except Exception:
                    pass
                return False

            def select_kategorie_second():
                try:
                    selector = "select[name='category'], select#category"
                    select_loc = page.locator(selector)
                    if select_loc.count() > 0 and select_loc.is_visible(timeout=500):
                        options_elements = select_loc.locator("option").all()
                        best_value = None
                        best_score = -1
                        best_label = ""
                        
                        title_lower = title.lower()
                        description_lower = description.lower()
                        
                        for opt in options_elements:
                            val = opt.get_attribute("value")
                            if not val or val == "" or val == "0":
                                continue
                            label = opt.inner_text().strip().lower()
                            
                            score = 0
                            if label in title_lower:
                                score += 10
                            if label in description_lower:
                                score += 2
                            
                            # Specifické Bazoš podkategorie
                            if "sekack" in val or "sekačk" in label:
                                if "sekačk" in title_lower or "vyžínač" in title_lower or "strunov" in title_lower:
                                    score += 20
                            if "drtic" in val or "drtič" in label:
                                if "drtič" in title_lower or "štěpkovač" in title_lower:
                                    score += 20
                            if "stol" in val or "stůl" in label:
                                if "stůl" in title_lower or "stoly" in title_lower:
                                    score += 20
                            if "židl" in label or "zidl" in val:
                                if "židl" in title_lower or "židle" in title_lower:
                                    score += 20
                                    
                            if score > best_score:
                                best_score = score
                                best_value = val
                                best_label = opt.inner_text().strip()
                        
                        if best_value and best_score > 0:
                            current_value = select_loc.evaluate("el => el.value")
                            if current_value != best_value:
                                print(f"  {Colors.BLUE}Volba kategorie '{best_label}'...{Colors.ENDC}")
                                select_loc.select_option(value=best_value)
                                time.sleep(0.5)
                                print(f"  {Colors.GREEN}✓ Kategorie vybrána: '{best_label}'{Colors.ENDC}")
                                return True
                except Exception:
                    pass
                return False

            def autofill_step1():
                try:
                    phone_input = page.locator("input[name='teloverit']")
                    if phone_input.is_visible():
                        phone_input.fill(user_config.get("phone", "777123456"))
                        terms_checkbox = page.locator("input[name='podminky']")
                        if terms_checkbox.is_visible() and not terms_checkbox.is_checked():
                            terms_checkbox.check()
                        print(f"\n{Colors.GREEN}[Krok 1] Telefonní číslo a souhlas vyplněny!{Colors.ENDC}")
                        
                        # Automaticky odešleme formulář krok 1
                        submit_btn = page.locator("input[type='submit']")
                        if submit_btn.count() > 0:
                            submit_btn.first.click()
                        else:
                            phone_input.press("Enter")
                            
                        print(f"  {Colors.GREEN}✓ Formulář ověření odeslán.{Colors.ENDC}")
                        return True
                except Exception:
                    pass
                return False

            step1_filled = False
            form_filled = False
            
            print(f"\n{Colors.BOLD}Sleduj okno prohlížeče a proveď SMS ověření (pokud je vyžadováno).{Colors.ENDC}")
            print("Jakmile se načte formulář, skript ho automaticky vyplní a nahraje fotky.")
            
            try:
                while not form_filled:
                    if page.is_closed():
                        print(f"\n{Colors.FAIL}Okno prohlížeče bylo zavřeno.{Colors.ENDC}")
                        break
                        
                    if not step1_filled:
                        step1_filled = autofill_step1()
                    
                    # Detekce formuláře inzerátu
                    nadpis_selectors = ["input[name='nadpis']", "#nadpis", "input[placeholder*='nadpis']"]
                    nadpis_found = False
                    for sel in nadpis_selectors:
                        try:
                            if page.locator(sel).is_visible(timeout=500):
                                nadpis_found = True
                                break
                        except Exception:
                            pass
 
                    if nadpis_found:
                        # 1. Nejprve zkusíme navolit Rubriku. Pokud to vyvolalo reload, cyklus pokračuje novou iterací
                        if select_rubrika_first():
                            continue
                            
                        # 2. Poté zkusíme navolit Kategorii
                        select_kategorie_second()
                        
                        # 3. Teprve nyní vyplníme textová pole
                        print(f"\n{Colors.GREEN}🎉 Formulář detekován! Začínám automaticky vyplňovat...{Colors.ENDC}")
                        find_and_fill("Nadpis", ["input[name='nadpis']", "#nadpis"], title)
                        find_and_fill("Popis", ["textarea[name='popis']", "#popis"], description)
                        find_and_fill("Cena", ["input[name='cena']", "#cena"], price)
                        find_and_fill("Jméno", ["input[name='jmeno']", "#jmeno"], user_config.get("name", "Tvoje Jméno"))
                        find_and_fill("Telefon", ["input[name='telefoni']", "#telefoni"], user_config.get("phone", "777123456"))
                        find_and_fill("E-mail", ["input[name='mail']", "input[name='email']", "input[type='email']", "#mail"], user_config.get("email", "tuj_email@example.com"))
                        find_and_fill("PSČ", ["input[name='lokalita']", "#lokalita", "input[name='psc']"], user_config.get("zip_code", "10000"))
                        find_and_fill("Heslo", ["input[name='heslo']", "#heslo"], password)
                        
                        # 4. Nahrání fotek
                        if photos:
                            print(f"{Colors.BLUE}Nahrávám {len(photos)} fotek...{Colors.ENDC}")
                            file_input_selectors = ["input[type='file']", "input[name='pfile[]']", "input[name*='file']"]
                            file_input_found = False
                            for sel in file_input_selectors:
                                try:
                                    loc = page.locator(sel)
                                    if loc.count() > 0:
                                        loc.first.set_input_files(photos)
                                        print(f"{Colors.GREEN}✓ Fotky úspěšně nahrány!{Colors.ENDC}")
                                        file_input_found = True
                                        break
                                except Exception:
                                    continue
                            if not file_input_found:
                                print(f"{Colors.WARNING}⚠️  Fotky se nepodařilo nahrát automaticky. Nahraj je ručně.{Colors.ENDC}")
                        
                        print(f"\n{Colors.HEADER}{Colors.BOLD}✅ HOTOVO! Všechno předvyplněno!{Colors.ENDC}")
                        print(f"{Colors.BOLD}👉 Zkontroluj inzerát v prohlížeči, ulož jej a zkopíruj si jeho URL.{Colors.ENDC}")
                        form_filled = True
                        break
                    
                    time.sleep(1)
                    
            except KeyboardInterrupt:
                print("\nKrok přerušen uživatelem.")
                return False
                
            if form_filled:
                if is_web:
                    new_url = ""
                else:
                    new_url = input(f"\n{Colors.BOLD}Zadej novou URL adresu inzerátu na Bazoši (pokud ji máš, jinak stiskni Enter): {Colors.ENDC}").strip()
                if new_url:
                    ad["url"] = new_url
                    ad["date_created"] = datetime.today().strftime('%Y-%m-%d')
                    print(f"{Colors.GREEN}✓ URL inzerátu uložena!{Colors.ENDC}")
                try:
                    context.storage_state(path=str(SESSION_STATE_PATH))
                except Exception:
                    pass
                return True
    return False

def run_playwright_action(*args, **kwargs):
    try:
        return _run_playwright_action_impl(*args, **kwargs)
    except Exception as e:
        print(f"\n{Colors.FAIL}Playwright operace byla přerušena nebo selhala: {e}{Colors.ENDC}")
        try:
            session_manager.close()
        except Exception:
            pass
        return False

# --- Znovuvystavení (Topování zdarma - smazat + vystavit znovu) ---
def cli_repost_listing(data, user_config):
    active = data.get("active_listings", [])
    if not active:
        print(f"{Colors.FAIL}Nemáš žádné aktivní inzeráty pro znovuvystavení!{Colors.ENDC}")
        return
        
    print(f"\n{Colors.HEADER}{Colors.BOLD}🔄 ZNOVUVYSTAVENÍ INZERÁTU (TOPOVÁNÍ ZDARMA){Colors.ENDC}")
    print(f"{Colors.BLUE}Tato funkce nejprve smaže stávající inzerát z Bazoše pomocí hesla a poté ihned spustí robota pro jeho nové vystavení.{Colors.ENDC}\n")
    
    for idx, ad in enumerate(active, 1):
        print(f"  [{idx}] {ad['title']} ({ad.get('url', 'Chybí odkaz')})")
        
    choice = input(f"\nZadej číslo inzerátu k znovuvystavení (1-{len(active)}): ").strip()
    try:
        idx = int(choice) - 1
        if idx < 0 or idx >= len(active):
            raise ValueError
        selected_ad = active[idx]
    except ValueError:
        print(f"{Colors.FAIL}Neplatná volba!{Colors.ENDC}")
        return
        
    confirm = input(f"Opravdu smazat a znovu vystavit '{selected_ad['title']}'? [y/N]: ").strip().lower()
    if confirm != 'y':
        print("Akce zrušena.")
        return
        
    # Společná funkce změny ceny při znovuvystavení
    change_price_opt = input(f"Chceš při této příležitosti rovnou změnit cenu inzerátu? (aktuálně: {selected_ad['price']} Kč) [y/N]: ").strip().lower()
    if change_price_opt == 'y':
        try:
            new_price = int(input("Zadej novou cenu v Kč: ").strip())
            selected_ad["price"] = new_price
            print(f"{Colors.GREEN}✓ Cena místně v databázi změněna na {new_price} Kč. Nová cena bude odeslána při vystavení.{Colors.ENDC}")
        except ValueError:
            print(f"{Colors.FAIL}Neplatná cena, ponechávám původní ({selected_ad['price']} Kč).{Colors.ENDC}")

        
    # Krok 1: Smazání
    if selected_ad.get("url"):
        print(f"\n{Colors.BLUE}Krok 1/2: Mažu inzerát z Bazoše...{Colors.ENDC}")
        run_playwright_action(selected_ad, user_config, action="delete")
    else:
        print(f"\n{Colors.WARNING}Inzerát nemá uloženou URL. Přeskakuji krok smazání a rovnou vystavuji...{Colors.ENDC}")
        
    # Krok 2: Nové vystavení
    print(f"\n{Colors.BLUE}Krok 2/2: Vystavuji inzerát znovu...{Colors.ENDC}")
    success = run_playwright_action(selected_ad, user_config, action="post")
    
    if success:
        selected_ad["date_created"] = datetime.today().strftime('%Y-%m-%d')
        save_listings(data)
        sync_to_onedrive(data)
        print(f"\n{Colors.GREEN}✓ Inzerát byl úspěšně znovuvystaven na Bazoši a uložen na OneDrive!{Colors.ENDC}")

# --- Hlavní spouštěcí funkce a menu ---
def main():
    print(f"\n{Colors.HEADER}{Colors.BOLD}🤖 Bazoš Automat v1.2{Colors.ENDC}")
    print(f"{Colors.BLUE}Komplexní správa inzerce, historie prodejů a synchronizace na OneDrive TERMS.{Colors.ENDC}\n")

    while True:
        data, user_config = load_data()
        
        print(f"\n{Colors.BOLD}⚙️  HLAVNÍ MENU AUTOMATU:{Colors.ENDC}")
        print(f"  [{Colors.GREEN}1{Colors.ENDC}] Poloautomaticky vystavit inzerát (Bazoš robot)")
        print(f"  [{Colors.GREEN}2{Colors.ENDC}] Přidat novou věc k inzerci (Průvodce)")
        print(f"  [{Colors.GREEN}3{Colors.ENDC}] Změnit cenu aktivního inzerátu")
        print(f"  [{Colors.GREEN}4{Colors.ENDC}] Znovuvystavit inzerát (Topovat zdarma)")
        print(f"  [{Colors.GREEN}5{Colors.ENDC}] Zaznamenat prodej věci (Přesunout do historie)")
        print(f"  [{Colors.GREEN}6{Colors.ENDC}] Zobrazit přehled inzerce v terminálu")
        print(f"  [{Colors.GREEN}7{Colors.ENDC}] Aktualizovat stavy a zobrazení inzerátů z Bazoše")
        print(f"  [{Colors.GREEN}8{Colors.ENDC}] Vynutit synchronizaci s OneDrivem (Excel)")
        print(f"  [{Colors.FAIL}q{Colors.ENDC}] Konec")
        
        choice = input(f"\n{Colors.BOLD}Zadej volbu (1-8, q): {Colors.ENDC}").strip()
        
        if choice.lower() == 'q':
            print("Nashledanou!")
            break
        elif choice == '1':
            active = data.get("active_listings", [])
            if not active:
                print(f"{Colors.FAIL}Žádné aktivní inzeráty v databázi! Přidej věc přes volbu 2.{Colors.ENDC}")
                continue
            print(f"\n{Colors.BOLD}Který inzerát chceš poloautomaticky vystavit?{Colors.ENDC}")
            for idx, ad in enumerate(active, 1):
                price_str = f"{ad['price']} Kč" if ad['price'] > 0 else "Zdarma"
                print(f"  [{Colors.GREEN}{idx}{Colors.ENDC}] {Colors.BOLD}{ad['title']}{Colors.ENDC} ({price_str})")
            
            ad_choice = input(f"Zadej číslo (1-{len(active)}) nebo Enter pro návrat: ").strip()
            if not ad_choice:
                continue
            try:
                ad_idx = int(ad_choice) - 1
                if ad_idx < 0 or ad_idx >= len(active):
                    raise ValueError
                selected_ad = active[ad_idx]
                run_playwright_action(selected_ad, user_config, action="post")
                save_listings(data)
                sync_to_onedrive(data)
            except ValueError:
                print(f"{Colors.FAIL}Neplatná volba!{Colors.ENDC}")
        elif choice == '2':
            cli_add_listing(data, user_config)
        elif choice == '3':
            cli_change_price(data, user_config, run_playwright_action)
        elif choice == '4':
            cli_repost_listing(data, user_config)
        elif choice == '5':
            cli_record_sale(data)
        elif choice == '6':
            display_listings_summary(data)
        elif choice == '7':
            cli_update_listings_from_bazos(data)
        elif choice == '8':
            sync_to_onedrive(data)
        else:
            print(f"{Colors.FAIL}Neplatná volba!{Colors.ENDC}")

if __name__ == "__main__":
    main()
