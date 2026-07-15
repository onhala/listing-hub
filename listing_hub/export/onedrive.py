import os
from pathlib import Path
from datetime import datetime
from listing_hub.core.db import get_all_listings

class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def sync_to_onedrive(listings=None):
    """
    Vygeneruje styled Excel tabulku z inzerátů a uloží ji do OneDrive složky TERMS.
    Pokud nedostane listings jako argument, načte je přímo z SQLite databáze.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            f"{Colors.FAIL}Knihovna openpyxl není nainstalována. "
            f"Spusťte: pip install openpyxl{Colors.ENDC}"
        )

    if listings is None:
        listings = get_all_listings()

    # Detekce export cesty
    # Podpora environment proměnné BAZOS_EXPORT_DIR z docker-compose
    export_env = os.environ.get("BAZOS_EXPORT_DIR")
    if export_env:
        onedrive_dir = Path(export_env)
    else:
        onedrive_dir = Path.home() / "Library/CloudStorage/OneDrive-Osobní"
        if not onedrive_dir.exists():
            onedrive_dir = Path.home() / "Desktop/Antigravity"

    onedrive_dir.mkdir(parents=True, exist_ok=True)
    file_path = onedrive_dir / "Inzerce - listing-hub.xlsx"
    
    wb = openpyxl.Workbook()
    
    # První list: Aktivní inzerce
    ws_active = wb.active
    ws_active.title = "Aktivní inzerce"
    
    # Druhý list: Historie prodejů
    ws_sold = wb.create_sheet(title="Historie prodejů")
    
    # Styly pro prémiový TERMS vzhled
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B2C82", end_color="4B2C82", fill_type="solid") # TERMS fialová
    title_font = Font(name=font_family, size=16, bold=True, color="4B2C82")
    data_font = Font(name=font_family, size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Rozdělíme inzeráty na aktivní a prodané
    active_listings = []
    sold_listings = []
    
    for item in listings:
        # Bazoš je primární fallback, pokud inzerát nemá specifické stavy, dáme ho do aktivních
        bazos_state = item.get("portal_states", {}).get("bazos", {})
        status = bazos_state.get("status", "Aktivní")
        
        if status == "Prodané" or item.get("notes") == "Prodané":
            sold_listings.append(item)
        else:
            active_listings.append(item)

    # Pomocná funkce pro nastavení listu
    def setup_sheet(ws, sheet_title, items, headers, is_sold=False):
        ws.views.sheetView[0].showGridLines = True
        
        # Titulní nadpis
        ws.merge_cells("A1:D1")
        ws["A1"] = sheet_title
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 30
        
        # Prázdný řádek 2
        ws.row_dimensions[2].height = 15
        
        # Hlavičky na řádku 3
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[3].height = 25
        
        current_row = 4
        for idx, ad in enumerate(items, 1):
            ws.row_dimensions[current_row].height = 22
            row_fill = PatternFill(start_color="F8F6FC" if idx % 2 == 0 else "FFFFFF", end_color="F8F6FC" if idx % 2 == 0 else "FFFFFF", fill_type="solid")
            
            bazos_state = ad.get("portal_states", {}).get("bazos", {})
            
            if not is_sold:
                days_old = ad.get("days_old", 0)
                date_created = ad.get("created_at", "")
                if date_created and days_old == 0:
                    try:
                        dt = datetime.strptime(date_created, "%Y-%m-%d")
                        days_old = (datetime.today() - dt).days
                    except Exception:
                        pass
                
                values = [
                    idx,
                    ad.get("title", ""),
                    ad.get("price", 0),
                    ad.get("created_at", ""),
                    days_old,
                    bazos_state.get("views", 0),
                    ad.get("condition", "Nezadáno"),
                    bazos_state.get("status", "Aktivní"),
                    bazos_state.get("url", ""),
                    ad.get("local_photos_dir", "")
                ]
            else:
                price = ad.get("price", 0)
                # Pokusíme se načíst sale_price ze zápisu, případně fallback na price
                sale_price = ad.get("sale_price", price)
                diff = sale_price - price
                values = [
                    idx,
                    ad.get("title", ""),
                    price,
                    sale_price,
                    diff,
                    ad.get("created_at", ""),
                    bazos_state.get("last_synced", "") or datetime.today().strftime("%Y-%m-%d"),
                    ad.get("condition", "Nezadáno"),
                    ad.get("notes", "")
                ]
                
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                
                if col_idx == 1:
                    cell.alignment = align_center
                elif not is_sold and col_idx == 3:
                    cell.alignment = align_right
                    cell.number_format = '#,##0" Kč"'
                elif is_sold and col_idx in [3, 4, 5]:
                    cell.alignment = align_right
                    cell.number_format = '#,##0" Kč"'
                elif not is_sold and col_idx in [4, 5, 6, 8]:
                    cell.alignment = align_center
                elif is_sold and col_idx in [6, 7]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
                if not is_sold and col_idx == 8:
                    status_val = bazos_state.get("status", "Aktivní")
                    if status_val == "Aktivní":
                        cell.fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, color="0F5132", bold=True)
                    elif status_val == "Expirováno":
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, color="842029", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, color="664D03", bold=True)
                    
                if not is_sold and col_idx == 9 and val:
                    cell.hyperlink = val
                    cell.font = Font(name=font_family, size=11, color="0D6EFD", underline="single")
                    
            current_row += 1
            
        if is_sold and len(items) > 0:
            ws.row_dimensions[current_row].height = 24
            ws.cell(row=current_row, column=2).value = "Celkem prodáno"
            ws.cell(row=current_row, column=2).font = Font(name=font_family, size=11, bold=True)
            ws.cell(row=current_row, column=2).alignment = align_right
            
            cell_orig = ws.cell(row=current_row, column=3)
            cell_orig.value = f"=SUM(C4:C{current_row-1})"
            cell_orig.font = Font(name=font_family, size=11, bold=True)
            cell_orig.alignment = align_right
            cell_orig.number_format = '#,##0" Kč"'
            cell_orig.border = thin_border
            
            cell_sale = ws.cell(row=current_row, column=4)
            cell_sale.value = f"=SUM(D4:D{current_row-1})"
            cell_sale.font = Font(name=font_family, size=11, bold=True, color="0F5132")
            cell_sale.alignment = align_right
            cell_sale.number_format = '#,##0" Kč"'
            cell_sale.border = thin_border
            cell_sale.fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
            
            cell_diff = ws.cell(row=current_row, column=5)
            cell_diff.value = f"=SUM(E4:E{current_row-1})"
            cell_diff.font = Font(name=font_family, size=11, bold=True)
            cell_diff.alignment = align_right
            cell_diff.number_format = '#,##0" Kč"'
            cell_diff.border = thin_border
            
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    setup_sheet(ws_active, "Aktivní inzerce", active_listings, ["ID", "Název věci", "Inzerovaná cena (Kč)", "Datum vystavení", "Stáří (dní)", "Zhlédnutí", "Stav věci", "Stav", "Odkaz na Bazoš", "Složka s fotkami"], is_sold=False)
    setup_sheet(ws_sold, "Historie prodejů", sold_listings, ["ID", "Název věci", "Původní cena (Kč)", "Prodejní cena (Kč)", "Rozdíl (Kč)", "Datum vystavení", "Datum prodeje", "Stav věci", "Poznámky"], is_sold=True)
    
    wb.save(file_path)
    print(f"{Colors.GREEN}✓ Tabulka Excel úspěšně synchronizována na OneDrive: {file_path}{Colors.ENDC}")
